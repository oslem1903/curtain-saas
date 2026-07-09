from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
dashboard = wb.active
dashboard.title = "Ana Sayfa"

# Başlık
dashboard['A1'] = "PERSONEL PUANTAJ YÖNETİM SİSTEMİ"
dashboard['A1'].font = Font(size=14, bold=True, color="FFFFFF")
dashboard['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
dashboard.merge_cells('A1:D1')
dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
dashboard.row_dimensions[1].height = 25

# Filtreler
dashboard['A3'] = "AYLAR SEÇ:"
dashboard['A3'].font = Font(bold=True)
dashboard['B3'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

dashboard['A4'] = "FİRMA SEÇ:"
dashboard['A4'].font = Font(bold=True)
dashboard['B4'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

dashboard['A5'] = "YIL SEÇ:"
dashboard['A5'].font = Font(bold=True)
dashboard['B5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Filtre açıklaması
dashboard['A7'] = "Seçilen Dönem Puantaj Listesi:"
dashboard['A7'].font = Font(bold=True, size=12)

# Tablo Başlıkları
headers = ["Personel Adı", "Pozisyon", "Maaşı", "İşe Giriş", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]
for col, header in enumerate(headers, 1):
    cell = dashboard.cell(row=8, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Kolon genişlikleri
dashboard.column_dimensions['A'].width = 18
dashboard.column_dimensions['B'].width = 15
dashboard.column_dimensions['C'].width = 12
dashboard.column_dimensions['D'].width = 12
for col in range(5, 15):
    dashboard.column_dimensions[get_column_letter(col)].width = 8

# 2. Personel Listesi
personel = wb.create_sheet("Personel Listesi")
personel_headers = ["Personel ID", "Adı Soyadı", "Firma", "Pozisyon", "Departman", "Maaşı (TL)", "İşe Giriş Tarihi", "Kimlik No", "Tel", "Mail", "Durum"]

for col, header in enumerate(personel_headers, 1):
    cell = personel.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')
personel.row_dimensions[1].height = 20

# Örnek veriler
example_data = [
    [1, "Ahmet YILMAZ", "Firma A", "Operatör", "Üretim", 5000, "01.01.2023", "12345678901", "05551234567", "ahmet@mail.com", "Aktif"],
    [2, "Fatma ŞAHİN", "Firma A", "Memur", "Yönetim", 4500, "15.03.2023", "98765432101", "05551111111", "fatma@mail.com", "Aktif"],
    [3, "Mehmet KAYA", "Firma B", "Şef", "Üretim", 6000, "20.02.2023", "55555555555", "05552222222", "mehmet@mail.com", "Aktif"],
]

for row_idx, row_data in enumerate(example_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        personel.cell(row=row_idx, column=col_idx).value = value

for col in range(1, 12):
    personel.column_dimensions[get_column_letter(col)].width = 14

# 3. Firmalar
firmalar = wb.create_sheet("Firmalar")
firma_headers = ["Firma ID", "Firma Adı", "Yetkili Kişi", "Telefon", "Mail", "Adres"]
for col, header in enumerate(firma_headers, 1):
    cell = firmalar.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

firma_data = [
    [1, "Firma A", "Ali Özdemir", "05551234567", "ali@firmaA.com", "İstanbul"],
    [2, "Firma B", "Zeynep Yıldız", "05559876543", "zeynep@firmaB.com", "Ankara"],
]

for row_idx, row_data in enumerate(firma_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        firmalar.cell(row=row_idx, column=col_idx).value = value

for col in range(1, 7):
    firmalar.column_dimensions[get_column_letter(col)].width = 16

# 4. Puantaj Şablonu
puantaj = wb.create_sheet("Puantaj Şablonu")
puantaj['A1'] = "PUANTAJ SAYFASI"
puantaj['A1'].font = Font(size=12, bold=True)
puantaj.merge_cells('A1:F1')

puantaj['A2'] = "Firma:"
puantaj['B2'] = "(Seçilecek)"
puantaj['A3'] = "Ay:"
puantaj['B3'] = "(Seçilecek)"
puantaj['A4'] = "Yıl:"
puantaj['B4'] = "(Seçilecek)"

puantaj['A6'] = "Gün"
puantaj['B6'] = "Personel Adı"
puantaj['C6'] = "Giriş Saati"
puantaj['D6'] = "Çıkış Saati"
puantaj['E6'] = "Çalışma Saati"
puantaj['F6'] = "Notlar"

for col in range(1, 7):
    puantaj.cell(row=6, column=col).font = Font(bold=True, color="FFFFFF")
    puantaj.cell(row=6, column=col).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

puantaj.column_dimensions['A'].width = 10
puantaj.column_dimensions['B'].width = 15
puantaj.column_dimensions['C'].width = 12
puantaj.column_dimensions['D'].width = 12
puantaj.column_dimensions['E'].width = 14
puantaj.column_dimensions['F'].width = 20

# Aylar referansı
aylar = wb.create_sheet("Aylar", 0)
aylar.column_dimensions['A'].hidden = True
ay_listesi = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
              "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
for idx, ay in enumerate(ay_listesi, 1):
    aylar.cell(row=idx, column=1).value = ay

# Yıllar referansı
yillar = wb.create_sheet("Yıllar")
yillar.column_dimensions['A'].hidden = True
for yil in range(2023, 2027):
    yillar.cell(row=yil-2022, column=1).value = yil

wb.save(r'C:\Users\ofis\Desktop\MAAŞ HESAP.xlsm')
print("✅ Başarılı! Excel dosyası oluşturuldu.")
